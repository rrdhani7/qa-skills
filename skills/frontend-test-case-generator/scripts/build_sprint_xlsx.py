#!/usr/bin/env python3
"""Write one Jira ticket's test cases into the sprint workbook.

The sprint workbook holds one tab per Jira ticket. This script rewrites only the
tab named after the ticket and leaves every other tab untouched — authoring the
.xlsx by hand is how sibling tickets get destroyed.

Usage:
    python3 build_sprint_xlsx.py cases.json [--vault PATH] [--dry-run]

Input JSON:
    {
      "sprint": "ES-176",
      "ticket": "ES-8348",
      "cases": [
        {
          "summary": "Scripter can add a question successfully",
          "priority": "High",
          "step_summary": "Given ...\nWhen ...\nThen ...",
          "test_data": "...",
          "expected_result": "...",
          "testing_phase": "Feature",
          "automation_status": "PLAN",
          "ground_truth": "PRD §4.3 — \"...\""
        }
      ]
    }

Case keys are the template's column headers in snake_case. `story_linkages`
defaults to the ticket id. `status` and `bug_links` must stay empty — titis-tcms
owns them.

Reads:  Templates/test_case_template.xlsx  (never written)
Writes: Drafts/test-cases/{sprint}.xlsx    (never test-cases/)
"""

from __future__ import annotations

import argparse
import copy
import json
import re
import shutil
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is not installed. Run: python3 -m pip install openpyxl")

TEMPLATE_REL = Path("Templates/test_case_template.xlsx")
OUT_DIR_REL = Path("Drafts/test-cases")
READONLY_REL = Path("test-cases")
LOCKED_COLUMNS = {"status", "bug_links"}
INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")


def die(msg: str) -> None:
    sys.exit(f"error: {msg}")


def header_key(header: str) -> str:
    return header.strip().lower().replace(" ", "_")


def resolve_vault(explicit: str | None) -> Path:
    if explicit:
        vault = Path(explicit).expanduser().resolve()
    else:
        # <vault>/.claude/skills/<skill>/scripts/build_sprint_xlsx.py
        vault = Path(__file__).resolve().parents[4]
    if not (vault / TEMPLATE_REL).is_file():
        die(f"{TEMPLATE_REL} not found under {vault}. Pass --vault.")
    return vault


def load_payload(path: Path) -> tuple[str, str, list[dict]]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        die(f"{path} is not valid JSON: {exc}")
    for field in ("sprint", "ticket", "cases"):
        if not payload.get(field):
            die(f"{path} is missing a non-empty '{field}'.")
    if not isinstance(payload["cases"], list):
        die("'cases' must be a list.")
    sprint = str(payload["sprint"]).strip()
    ticket = str(payload["ticket"]).strip()
    if INVALID_SHEET_CHARS.search(ticket) or len(ticket) > 31:
        die(f"ticket '{ticket}' is not a usable sheet name (max 31 chars, no \\/*?:[]).")
    return sprint, ticket, payload["cases"]


def read_template(path: Path) -> tuple[list[str], dict, list, object]:
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    if not all(headers):
        die("template header row has an empty cell.")
    widths = {k: v.width for k, v in ws.column_dimensions.items()}
    header_styles = [
        (copy.copy(c.font), copy.copy(c.fill), copy.copy(c.alignment), copy.copy(c.border))
        for c in ws[1]
    ]
    return headers, widths, header_styles, ws[1][0].alignment


def validate_cases(cases: list[dict], headers: list[str], ticket: str) -> list[dict]:
    known = {header_key(h) for h in headers}
    cleaned = []
    for i, case in enumerate(cases, 1):
        if not isinstance(case, dict):
            die(f"case {i} is not an object.")
        unknown = set(case) - known
        if unknown:
            die(f"case {i} has unknown keys {sorted(unknown)}. Known: {sorted(known)}")
        for locked in LOCKED_COLUMNS & set(case):
            if str(case[locked]).strip():
                die(f"case {i} sets '{locked}'; titis-tcms owns that column — leave it empty.")
        if not str(case.get("summary", "")).strip():
            die(f"case {i} has no summary.")
        if not str(case.get("ground_truth", "")).strip():
            die(f"case {i} has no ground_truth; every case needs a source ref.")
        case = dict(case)
        case.setdefault("story_linkages", ticket)
        cleaned.append(case)
    return cleaned


def write_sheet(ws, headers, widths, header_styles, cases) -> None:
    if ws.max_row:
        ws.delete_rows(1, ws.max_row)
    ws.append(headers)
    for cell, (font, fill, alignment, border) in zip(ws[1], header_styles):
        cell.font, cell.fill, cell.alignment, cell.border = font, fill, alignment, border
    body_alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
    for case in cases:
        ws.append([case.get(header_key(h), "") or "" for h in headers])
        for cell in ws[ws.max_row]:
            cell.alignment = body_alignment
    for col, width in widths.items():
        if width:
            ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("cases_json", type=Path)
    parser.add_argument("--vault", help="vault root (default: inferred from script path)")
    parser.add_argument("--dry-run", action="store_true", help="report the plan, write nothing")
    args = parser.parse_args()

    if not args.cases_json.is_file():
        die(f"{args.cases_json} not found.")

    vault = resolve_vault(args.vault)
    template = vault / TEMPLATE_REL
    out_dir = vault / OUT_DIR_REL
    sprint, ticket, raw_cases = load_payload(args.cases_json)
    out = out_dir / f"{sprint}.xlsx"

    if (vault / READONLY_REL) in out.resolve().parents:
        die("refusing to write under test-cases/ — that folder is the read-only tcms sync target.")

    headers, widths, header_styles, _ = read_template(template)
    cases = validate_cases(raw_cases, headers, ticket)

    existed = out.is_file()
    before = openpyxl.load_workbook(out).sheetnames if existed else []
    print(f"sprint file : {out.relative_to(vault)} ({'exists' if existed else 'new'})")
    print(f"sheets before: {before or '—'}")
    print(f"ticket tab  : {ticket} ({'replaced' if ticket in before else 'added'}), {len(cases)} case(s)")

    if args.dry_run:
        print("dry run — nothing written.")
        return

    out_dir.mkdir(parents=True, exist_ok=True)
    if existed:
        wb = openpyxl.load_workbook(out)
        position = wb.sheetnames.index(ticket) if ticket in wb.sheetnames else None
        if position is not None:
            del wb[ticket]
        ws = wb.create_sheet(ticket, position)  # None appends; an int keeps the tab order
    else:
        shutil.copyfile(template, out)
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        ws.title = ticket

    write_sheet(ws, headers, widths, header_styles, cases)
    wb.save(out)

    after = openpyxl.load_workbook(out).sheetnames
    print(f"sheets after : {after}")
    lost = [s for s in before if s != ticket and s not in after]
    if lost:
        die(f"sheets disappeared: {lost} — restore the file from git before continuing.")
    print("ok")


if __name__ == "__main__":
    main()
